from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.core.paginator import Paginator
from .models import Miembro, Membresia
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io

@login_required
def reporte_miembros_activos_inactivos(request):
    hoy = timezone.now().date()
    # --- FILTRO Y BUSQUEDA ---
    busqueda_activos = request.GET.get('busqueda_activos', '').strip()
    busqueda_inactivos = request.GET.get('busqueda_inactivos', '').strip()

    miembros_activos_qs = Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct()
    miembros_inactivos_qs = Miembro.objects.exclude(id__in=miembros_activos_qs.values_list('id', flat=True))

    if busqueda_activos:
        miembros_activos_qs = miembros_activos_qs.filter(
            nombre__icontains=busqueda_activos
        ) | miembros_activos_qs.filter(
            apellido__icontains=busqueda_activos
        ) | miembros_activos_qs.filter(
            cedula__icontains=busqueda_activos
        )
    if busqueda_inactivos:
        miembros_inactivos_qs = miembros_inactivos_qs.filter(
            nombre__icontains=busqueda_inactivos
        ) | miembros_inactivos_qs.filter(
            apellido__icontains=busqueda_inactivos
        ) | miembros_inactivos_qs.filter(
            cedula__icontains=busqueda_inactivos
        )

    # --- PAGINACION ---
    paginador_activos = Paginator(miembros_activos_qs, 10)
    paginador_inactivos = Paginator(miembros_inactivos_qs, 10)
    page_num_activos = request.GET.get('page_activos', 1)
    page_num_inactivos = request.GET.get('page_inactivos', 1)
    page_activos = paginador_activos.get_page(page_num_activos)
    page_inactivos = paginador_inactivos.get_page(page_num_inactivos)

    context = {
        'page_activos': page_activos,
        'page_inactivos': page_inactivos,
        'busqueda_activos': busqueda_activos,
        'busqueda_inactivos': busqueda_inactivos,
        'total_activos': miembros_activos_qs.count(),
        'total_inactivos': miembros_inactivos_qs.count(),
        'total_activos_all': Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct().count(),
        'total_inactivos_all': Miembro.objects.exclude(id__in=miembros_activos_qs.values_list('id', flat=True)).count(),
    }
    return render(request, 'app_gimnasio/reporte_miembros_activos_inactivos.html', context)

@login_required
def exportar_miembros_excel(request):
    hoy = timezone.now().date()
    tipo = request.GET.get('tipo', 'activos')
    busqueda = request.GET.get('busqueda', '').strip()
    if tipo == 'activos':
        miembros_qs = Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct()
    else:
        miembros_activos = Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct()
        miembros_qs = Miembro.objects.exclude(id__in=miembros_activos.values_list('id', flat=True))
    if busqueda:
        miembros_qs = miembros_qs.filter(
            nombre__icontains=busqueda
        ) | miembros_qs.filter(
            apellido__icontains=busqueda
        ) | miembros_qs.filter(
            cedula__icontains=busqueda
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Miembros'
    ws.append(['Nombre', 'Apellido', 'Cédula', 'Estado'])
    for m in miembros_qs:
        ws.append([
            m.nombre,
            m.apellido,
            m.cedula,
            'Activo' if tipo == 'activos' else 'Inactivo'
        ])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"miembros_{tipo}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
def exportar_miembros_pdf(request):
    hoy = timezone.now().date()
    tipo = request.GET.get('tipo', 'activos')
    busqueda = request.GET.get('busqueda', '').strip()
    if tipo == 'activos':
        miembros_qs = Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct()
    else:
        miembros_activos = Miembro.objects.filter(membresia__activa=True, membresia__fecha_vencimiento__gte=hoy).distinct()
        miembros_qs = Miembro.objects.exclude(id__in=miembros_activos.values_list('id', flat=True))
    if busqueda:
        miembros_qs = miembros_qs.filter(
            nombre__icontains=busqueda
        ) | miembros_qs.filter(
            apellido__icontains=busqueda
        ) | miembros_qs.filter(
            cedula__icontains=busqueda
        )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    title = Paragraph(f"Miembros {tipo.title()}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    data = [['Nombre', 'Apellido', 'Cédula']]
    for m in miembros_qs:
        data.append([m.nombre, m.apellido, m.cedula])
    if len(data) == 1:
        data.append(['No hay miembros para mostrar.', '', ''])
    table = Table(data, colWidths=[150, 150, 120])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#223A5E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6EC6F1')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#223A5E')),
    ]))
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"miembros_{tipo}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
